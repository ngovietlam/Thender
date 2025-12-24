from openpyxl import load_workbook


class PartnerService:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self._cache = None

    def get_all(self):
        if self._cache is None:
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb["PARTNERS"]
            self._cache = [
                r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]
            ]
        return self._cache

    def add_if_not_exists(self, partner_name):
        partner_name = partner_name.strip()
        if not partner_name:
            return

        partners = self.get_all()
        if partner_name in partners:
            return

        wb = load_workbook(self.excel_file)
        ws = wb["PARTNERS"]

        ws.append([partner_name])
        wb.save(self.excel_file)

        self._cache.append(partner_name)
