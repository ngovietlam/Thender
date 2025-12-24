from openpyxl import load_workbook


class ExcelService:
    def __init__(self, excel_file, default_sheet):
        self.excel_file = excel_file
        self.default_sheet = default_sheet

    def load(self):
        """Load workbook and default sheet."""
        wb = load_workbook(self.excel_file)
        return wb, wb[self.default_sheet]

    def load_sheet(self, sheet_name):
        """Load workbook and any sheet by name."""
        wb = load_workbook(self.excel_file)
        return wb, wb[sheet_name]

    def find_row_by_code(self, ws, code):
        """Find a row by code in column B."""
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value == code:
                return r
        return None

    def find_next_row(self, ws):
        """Find next empty row based on column A."""
        for r in range(ws.max_row, 1, -1):
            if ws.cell(r, 1).value not in (None, ""):
                return r + 1
        return 2
