from openpyxl import load_workbook
import os, traceback
p = r"f:\Admin\Application\NhapLieu\data.xlsx"
print('path:', p)
print('exists:', os.path.exists(p))
if not os.path.exists(p):
    print('Tệp không tồn tại ở vị trí trên.')
else:
    try:
        wb = load_workbook(p, data_only=True)
        print('sheets:', wb.sheetnames)
    except Exception:
        traceback.print_exc()
